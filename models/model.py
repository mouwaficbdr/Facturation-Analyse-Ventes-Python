import pandas as p
from openpyxl import load_workbook

class Model():
    __file_dir='data/'

# Fontion de création des entrées dans les fichiers
    def _createData(self, file, code_root, data):
        # Ouverture du fichier
        file_path=self.__file_dir + f"{file}"
        file_=load_workbook(file_path)
        new_file_=file_.active

        # Insertion des données dans le fichier apres génération du code unique
        if new_file_!= None :
            index=new_file_.max_row
            data.insert(0,self.__generateCode(code_root,index))
            new_file_.append(data)
            file_.save(file_path)
            file_.close

# Retourne toutes les entrées du tableau du fichier 
    def _getAllData(self, file):
        return p.DataFrame(p.read_excel(self.__file_dir + f"{file}")).to_dict(orient="records")

# Fontion de génération du code unique de l'entrée
    def __generateCode(self, code_root , index):
        if index < 10 :
            return code_root + '00' + f"{index}"
        elif index < 100 :
            return code_root + '0' + f"{index}"
        else :
            return code_root + f"{index}"
